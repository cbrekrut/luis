import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = 'итс 1313 0000 000.xlsx'
output_file = input_file.split(' ')[0]
def extract_filtered_labor_data(input_file):
    print('-'*6,'exract','-'*6)
    # Load the Excel file
    df = pd.read_excel(input_file, sheet_name='Расчет')
    
    # Find the row where "Наименование работ" appears
    work_name_col_index = df[df.apply(lambda row: row.astype(str).str.contains("Наименование работ", case=False).any(), axis=1)].index[0]
    
    # Find the row with "услуги субподрядчиков (сторонние организации)"
    end_row_index = df[df.apply(lambda row: row.astype(str).str.contains("услуги субподрядчиков", case=False).any(), axis=1)].index[0]

    start_col_index = df.columns.get_loc('ПРОЕКТ')  # One column to the left of 'Наименование работ'
    end_col_index = df.columns.get_loc('Unnamed: 2') + 5  # Five columns to the right

    extracted_data = df.iloc[work_name_col_index+1:end_row_index, start_col_index:end_col_index+1]
    
    filtered_data = extracted_data[extracted_data.iloc[:, 0].astype(str).str.match(r'^\d+')]
    # Сбросим индексы после удаления строк
    filtered_data = filtered_data.reset_index(drop=True)
    return filtered_data

def uzel(input_file):
    print('-'*6,'uzel','-'*6)
    xls = pd.ExcelFile(input_file)
    # Загрузка данных с листа "Расчет"
    df = pd.read_excel(xls, sheet_name='Расчет')
    # Поиск фразы "ПОУЗЛОВОЙ РАСЧЕТ СЕБЕСТОИМОСТИ"
    search_phrase = "ПОУЗЛОВОЙ РАСЧЕТ СЕБЕСТОИМОСТИ"
    mask = df.apply(lambda row: row.astype(str).str.contains(search_phrase, case=False).any(), axis=1)

    # Найдем индекс строки с фразой
    if mask.any():
        start_row = mask[mask].index[0] + 1
        resulting_table = df.iloc[start_row:].reset_index(drop=True)
        
        # Фильтрация строк по столбцу "ПРОЕКТ"
        def is_digit(value):
            try:
                float_value = float(value)
                return float_value.is_integer() 
            except (ValueError, TypeError):
                return False

        valid_indices = []
        for idx in resulting_table.index:
            if is_digit(resulting_table.at[idx, 'ПРОЕКТ']):
                valid_indices.append(idx)
                if idx + 1 < len(resulting_table):
                    valid_indices.append(idx + 1)
                if idx + 2 < len(resulting_table):
                    valid_indices.append(idx + 2)

        filtered_table = resulting_table.loc[valid_indices].reset_index(drop=True)

        row = 0
        while row < len(filtered_table):
            if filtered_table.iloc[row, 3] == 0:
                filtered_table = filtered_table.drop(filtered_table.index[row:])
                break
            row += 1
        # Сбросим индексы после удаления строк
        filtered_table = filtered_table.reset_index(drop=True)
        return filtered_table
    else:
        print("Фраза не найдена.")


file_primer_tablicy33 = 'Пример таблицы33.xlsx'

# Чтение исходных файлов
df_naimenovanie_rabot = extract_filtered_labor_data(input_file)
df_primer_tablicy33 = pd.read_excel(file_primer_tablicy33)
df_final_filtered_table = uzel(input_file)
# Создаем копию таблицы
df_updated = df_primer_tablicy33.copy()

# Очистка столбцов C и E начиная с первой строки
df_updated.iloc[:, 2] = np.nan  # Используем numpy.nan вместо pd.NA для столбца C (3-й столбец)
df_updated.iloc[:, 4] = np.nan  # Используем numpy.nan вместо pd.NA для столбца E (5-й столбец)

# Определяем минимальное количество строк, чтобы избежать ошибок
min_len = min(len(df_naimenovanie_rabot), len(df_primer_tablicy33))
df_updated.iloc[:, 7] = np.nan
# Заполняем очищенные столбцы новыми данными
df_updated.iloc[:min_len, 2] = df_naimenovanie_rabot.iloc[:min_len, 1].values  # Перенос данных из столбца B в C
df_updated.iloc[:min_len, 4] = df_naimenovanie_rabot.iloc[:min_len, 3].values  # Перенос данных из столбца D в E
df_updated.iloc[:min_len, 6] = df_naimenovanie_rabot.iloc[:min_len, 6].values  # Перенос данных из столбца G в G
df_updated.iloc[:min_len, 7] = df_naimenovanie_rabot.iloc[:min_len, 6].values  # Перенос данных из столбца G в G
df_updated.iloc[:min_len, 8] = df_naimenovanie_rabot.iloc[:min_len, 4].values  # Перенос данных из столбца G в G
df_updated = df_updated[~((df_updated.iloc[:, 6] == 0) | (df_updated.iloc[:, 6].isna()))]
df_updated = df_updated[~((df_updated.iloc[:, 4] == 0) | (df_updated.iloc[:, 4].isna()))]
# Разделение значения из столбца G на данные из столбца E с проверкой на деление на ноль
df_updated.iloc[:min_len, 6] = np.where(
    df_updated.iloc[:min_len, 4] != 0,  # Проверяем, что в столбце E нет нуля
    (df_updated.iloc[:min_len, 6] / df_updated.iloc[:min_len, 4] / 2).round(2),  # Делим значения
    np.nan  # Если в столбце E ноль, записываем nan
)
df_updated.iloc[:min_len, 7] = np.where(
    df_updated.iloc[:min_len, 4] != 0,  # Проверяем, что в столбце E нет нуля
    (df_updated.iloc[:min_len, 7] / df_updated.iloc[:min_len, 4]).round(2),  # Делим значения
    np.nan  # Если в столбце E ноль, записываем nan
)

# Удаление строк, где в столбце E (4-й столбец) значение 0 или пусто
df_updated = df_updated[~((df_updated.iloc[:, 4] == 0) | (df_updated.iloc[:, 4].isna()))]

# Теперь добавляем данные из final_filtered_table
# Берем данные из столбца D только из строк 3, 6, 9 и т.д.
rows_to_extract = range(1, len(df_final_filtered_table), 3)  # 3-я строка — это индекс 2
rows_to_extract1 = range(2, len(df_final_filtered_table), 3)  # 3-я строка — это индекс 2
values_to_add = df_final_filtered_table.iloc[rows_to_extract, 3].values
values_to_add1 = df_final_filtered_table.iloc[rows_to_extract1, 11].values  # Столбец L — это индекс 11
values_to_add2 = df_final_filtered_table.iloc[rows_to_extract1, 10].values  # Столбец K — это индекс 10
# Найдем текущее количество строк в df_updated
current_length = len(df_updated)
# Обеспечим наличие достаточного количества строк для добавления новых данных
new_rows = pd.DataFrame(np.nan, index=range(len(values_to_add)), columns=df_updated.columns)
df_updated = pd.concat([df_updated, new_rows], ignore_index=True)
# Добавляем эти значения в столбец E (5-й столбец) после существующих данных
df_updated.iloc[current_length:current_length + len(values_to_add), 4] = values_to_add
# Добавляем эти значения в столбец G (7-й столбец) после существующих данных
df_updated.iloc[current_length:current_length + len(values_to_add1), 6] = np.round(np.array(values_to_add1, dtype=float) / 2, 2)
# Добавляем эти значения в столбец H (8-й столбец) после существующих данных
df_updated.iloc[current_length:current_length + len(values_to_add1), 7] = np.round(np.array(values_to_add1, dtype=float), 2)
df_updated.iloc[current_length:current_length + len(values_to_add), 8] = values_to_add2
# Выполняем деление значений столбцов H и G на столбец E только для добавленных строк
df_updated.iloc[current_length:current_length + len(values_to_add), 6] = np.where(
    df_updated.iloc[current_length:current_length + len(values_to_add), 4] != 0,
    (df_updated.iloc[current_length:current_length + len(values_to_add), 6] / df_updated.iloc[current_length:current_length + len(values_to_add), 4]).round(2),
    np.nan
)
df_updated.iloc[current_length:current_length + len(values_to_add), 7] = np.where(
    df_updated.iloc[current_length:current_length + len(values_to_add), 4] != 0,
    (df_updated.iloc[current_length:current_length + len(values_to_add), 7] / df_updated.iloc[current_length:current_length + len(values_to_add), 4]).round(2),
    np.nan
)
df_updated.iloc[current_length:current_length + len(values_to_add), 8] = np.where(
    df_updated.iloc[current_length:current_length + len(values_to_add), 4] != 0,
    (df_updated.iloc[current_length:current_length + len(values_to_add), 8] / df_updated.iloc[current_length:current_length + len(values_to_add), 4]).round(1),
    np.nan
)
# Заполняем столбец C (3-й столбец) фразами "Сварка - 1", "Сварка - 2" и т.д.
df_updated.iloc[current_length:current_length + len(values_to_add), 2] = [f"{output_file} Сварка - {i+1}" for i in range(len(values_to_add))]
# Заполняем столбец B (2-й столбец) значением "Сварочный цех"
df_updated.iloc[current_length:current_length + len(values_to_add), 1] = "Сварочный цех"



# Заполняем столбец A (1-й столбец) названием выходного файла
df_updated.iloc[:, 0] = output_file
df_updated.iloc[:, 3] = 0
df_updated.iloc[:, 5] = 'серия'

df_updated.iloc[:, 9] = 0

# Добавляем ваш код здесь
# Обновляем столбец B на основе значений в столбце C
for i in range(len(df_updated)):
    cell_value = str(df_updated.iloc[i, 2]) # Столбец C
    if i == (len(df_updated)-1):
        df_updated.iloc[i, 1] = 'Сборочно-упаковочный цех' # Столбец B
    if 'Сварка сетки' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'  # Столбец B
    elif 'Крепеж' in cell_value:
        df_updated.iloc[i, 1] = 'Сборочно-упаковочный цех'
    elif 'Покрасочые работы' in cell_value:
        df_updated.iloc[i, 1] = 'Покрасочный цех'
        df_updated.iloc[i, 6] = 70   
    elif 'Распил на ленточной пиле прямой рез' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Распил на ленточной пиле косой рез' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Резка листа ПВЛ' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Резка трубок ПВХ' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Фрезеровка' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Сверление' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Гибка труб' in cell_value:
        df_updated.iloc[i, 1] = 'Заготовительный цех'
    elif 'Наклейка' in cell_value:
        df_updated.iloc[i, 1] = 'Сборочно-упаковочный цех'
    elif 'Маркировка' in cell_value:
        df_updated.iloc[i, 1] = 'Сборочно-упаковочный цех'
    elif 'Погрузочно-разгрузочные работы' in cell_value:
        df_updated.iloc[i, 1] = 'Сборочно-упаковочный цех'
        df_updated.iloc[i, 3] = df_updated.iloc[i, 4]/2
        df_updated.iloc[i, 4] = df_updated.iloc[i, 4]/2
    elif 'Упаковочные работы' in cell_value:
        df_updated.iloc[i, 1] = 'Сборочно-упаковочный цех'
        df_updated.iloc[i, 3] = df_updated.iloc[i, 4]
                
# Заполнение пустых строк в столбце I нулями
df_updated.iloc[:, 8] = df_updated.iloc[:, 8].fillna(0)

df_filtered = df_updated[df_updated.iloc[:, 1] == 'Сборочно-упаковочный цех']

# # Шаг 2: Если есть строки с "Сборочно-упаковочный цех"
if not df_filtered.empty:
    # Выполняем умножение значений в столбцах E и G
    multiplied_g = df_filtered.apply(lambda row: row[4] * row[6], axis=1)
    # Выполняем умножение значений в столбцах E и H
    multiplied_h = df_filtered.apply(lambda row: row[4] * row[7], axis=1)
    
    # Суммируем результаты умножения для столбцов G и H
    sum_g = multiplied_g.sum()  # Итоговая сумма для столбца G (7-й столбец)
    sum_h = multiplied_h.sum()  # Итоговая сумма для столбца H (8-й столбец)

    # Создаем новую строку с итоговыми значениями
    final_row = pd.Series([
        output_file,  # Столбец A (название файла)
        'Сборочно-упаковочный цех',  # Столбец B
        'Финальная сборка',  # Столбец C (текст "Финальная сборка")
        0,  # Столбец D
        1,  # Столбец E (значение 1)
        'серия',  # Столбец F
        sum_g,  # Столбец G (сумма значений)
        sum_h,  # Столбец H (сумма значений)
        0,  # Столбец I (оставляем 0)
        0  # Столбец I (оставляем 0)
    ], index=df_updated.columns)

    # Шаг 3: Удаляем все строки с "Сборочно-упаковочный цех"
    df_updated = df_updated[df_updated.iloc[:, 1] != 'Сборочно-упаковочный цех']

    # Шаг 4: Добавляем новую строку в конец таблицы
    df_updated = df_updated._append(final_row, ignore_index=True)

df = pd.read_excel(input_file, sheet_name='Расчет')
d2_value = df.iloc[0,3]  # Это значение из ячейки D2

df_updated.iloc[:, 4] = df_updated.iloc[:, 4] * d2_value
output_file += ' Наряд'

# Сохранение обновленной таблицы в файл
df_updated.to_excel(f'{output_file}.xlsx', index=False)

# Открываем файл с помощью openpyxl для изменения ширины столбцов
wb = load_workbook(f'{output_file}.xlsx')
ws = wb.active

# Устанавливаем ширину столбцов по максимальному значению
for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter  # Получаем букву столбца
    for cell in column_cells:
        try:
            # Вычисляем длину содержимого ячейки
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    # Устанавливаем ширину столбца
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Сохраняем изменения в файл
wb.save(f'{output_file}.xlsx')

print("Success")


#TODO 
# Исправление сварка кол-во узлов не сумировать
# Цена та что в ячейке
# Сборочно упаковочные работы первести в финальную сборку 